from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parent
OUTPUT_FILE = ROOT / "College_Management_System_Website_Test_Cases_And_Scenarios.xlsx"


THIN = Side(style="thin", color="C7D3E0")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="DCEAF7")
ALT_FILL = PatternFill("solid", fgColor="F7FAFC")
TITLE_FILL = PatternFill("solid", fgColor="EAF3FB")


def test_rows():
    return [
        ("Public Visitor", "Website Access", "Public Navigation", "TC-PUB-001", "Verify public website opens successfully", "Open the website in a browser", "Landing page should load with public navigation and main action buttons", "High"),
        ("Public Visitor", "Website Access", "Public Navigation", "TC-PUB-002", "Verify public pages are accessible without login", "Visit Home, About, and Contact pages", "Each public page should load correctly without login", "Medium"),
        ("Public Visitor", "Registration", "User Registration", "TC-PUB-003", "Verify register page opens", "Click Apply Now or Register", "Registration page should open", "High"),
        ("Public Visitor", "Registration", "Student Registration", "TC-PUB-004", "Verify student registration flow", "Enter full name, email, password, choose Student, add roll number, then register", "Student account should be created and redirected to student dashboard", "High"),
        ("Public Visitor", "Registration", "Faculty Registration", "TC-PUB-005", "Verify faculty registration flow", "Enter full name, email, password, choose Faculty, add employee ID, then register", "Faculty account should be created and redirected to faculty dashboard", "Medium"),
        ("Public Visitor", "Registration", "Admin Registration", "TC-PUB-006", "Verify admin registration flow", "Enter full name, email, password, choose Admin, then register", "Admin account should be created and redirected to admin dashboard", "Low"),
        ("All Roles", "Login", "Authentication", "TC-LOGIN-001", "Verify login page opens", "Open login screen", "Role selection cards should be visible", "High"),
        ("All Roles", "Login", "Student Login", "TC-LOGIN-002", "Verify student can log in", "Choose Student role, enter student credentials, click Sign In", "User should open /student/dashboard", "High"),
        ("All Roles", "Login", "Faculty Login", "TC-LOGIN-003", "Verify faculty can log in", "Choose Faculty role, choose Faculty Login, enter credentials, click Sign In", "User should open /faculty/dashboard", "High"),
        ("All Roles", "Login", "HOD Login", "TC-LOGIN-004", "Verify HOD can log in", "Choose Faculty role, choose HOD Login, enter credentials, click Sign In", "User should open /hod/dashboard", "High"),
        ("All Roles", "Login", "Admin Login", "TC-LOGIN-005", "Verify admin can log in", "Choose Admin role, enter admin credentials, click Sign In", "User should open /admin/dashboard", "High"),
        ("Student", "Dashboard", "Student Dashboard", "TC-STU-001", "Verify student dashboard loads", "Open student dashboard after login", "Dashboard cards, quick actions, notifications, upcoming exams, and attendance overview should appear", "High"),
        ("Student", "Profile", "Student Profile", "TC-STU-002", "Verify student profile page loads", "Open Profile from sidebar", "Student profile page should load", "Medium"),
        ("Student", "Attendance", "Student Attendance", "TC-STU-003", "Verify student attendance page loads", "Open Attendance from sidebar", "Attendance page should show subject-wise or overall attendance information", "High"),
        ("Student", "Timetable", "Student Timetable", "TC-STU-004", "Verify student timetable loads", "Open Timetable from sidebar", "Class schedule should load", "High"),
        ("Student", "Results", "Student Results", "TC-STU-005", "Verify student can view results", "Open Results from sidebar", "Published result data should load", "High"),
        ("Student", "Fees", "Student Fees", "TC-STU-006", "Verify student fees page loads", "Open Fees from sidebar", "Fee due or payment records should be visible", "High"),
        ("Student", "Notifications", "Student Notifications", "TC-STU-007", "Verify student notifications load", "Open Notifications from sidebar", "Student notices should be visible", "Medium"),
        ("Student", "Placements", "Student Placements", "TC-STU-008", "Verify student placements page loads", "Open Placements from sidebar", "Placement drives or applications should be visible", "Medium"),
        ("Student", "Library", "Student Library", "TC-STU-009", "Verify student library page loads", "Open Library from sidebar", "Library page should load with issued books or records", "Medium"),
        ("Student", "Certificates", "Student Certificates", "TC-STU-010", "Verify student certificates page loads", "Open Certificates from sidebar", "Certificate list or download options should appear", "Medium"),
        ("Student", "Hostel", "Student Hostel", "TC-STU-011", "Verify student hostel page loads", "Open Hostel from sidebar", "Hostel information should load", "Low"),
        ("Student", "Transport", "Student Transport", "TC-STU-012", "Verify student transport page loads", "Open Transport from sidebar", "Transport information should load", "Low"),
        ("Student", "Study Materials", "Student Materials", "TC-STU-013", "Verify student materials page loads", "Open Study Materials from sidebar", "Study material page should load", "Medium"),
        ("Student", "Face ID", "Face ID Registration", "TC-STU-014", "Verify student can access face ID registration", "Open Face ID Register page and perform registration flow", "Face registration workflow should complete without error", "Low"),
        ("Student", "Complaints", "Student Complaint", "TC-STU-015", "Verify student can submit complaint", "Open Complaints page and submit a complaint", "Complaint should be saved and appear in records", "Medium"),
        ("Student", "Service Request", "Student Service Request", "TC-STU-016", "Verify student can create service request", "Open Service Request page and submit a request", "Request should be created", "Medium"),
        ("Student", "Request Status", "Student Request Tracking", "TC-STU-017", "Verify student can track request status", "Open Request Status page", "Existing request status should be visible", "Medium"),
        ("Student", "Documents", "Student Documents", "TC-STU-018", "Verify student can upload document", "Open Documents page and upload a file", "Document should be uploaded and listed", "Medium"),
        ("Student", "Scholarship", "Student Scholarship", "TC-STU-019", "Verify scholarship page loads", "Open Scholarship page from sidebar", "Scholarship page should load", "Low"),
        ("Student", "Events", "Student Events", "TC-STU-020", "Verify campus events page loads", "Open Events page from sidebar", "Campus events page should load", "Low"),
        ("Faculty", "Dashboard", "Faculty Dashboard", "TC-FAC-001", "Verify faculty dashboard loads", "Login as faculty and open dashboard", "Faculty dashboard should load", "High"),
        ("Faculty", "Profile", "Faculty Profile", "TC-FAC-002", "Verify faculty profile loads", "Open Profile from sidebar", "Faculty profile should load", "Medium"),
        ("Faculty", "Timetable", "Faculty Timetable", "TC-FAC-003", "Verify faculty timetable loads", "Open Timetable from sidebar", "Faculty timetable should be visible", "High"),
        ("Faculty", "Notifications", "Faculty Notifications", "TC-FAC-004", "Verify faculty notifications load", "Open Notifications from sidebar", "Faculty notifications should load", "Medium"),
        ("Faculty", "Attendance", "Attendance Management", "TC-FAC-005", "Verify faculty can mark attendance", "Open Mark Attendance, choose course and date, mark statuses, submit", "Attendance should save and attendance report should refresh", "High"),
        ("Faculty", "Face Attendance", "Face Attendance", "TC-FAC-006", "Verify faculty face attendance workflow", "Open Face Attendance page and perform attendance workflow", "Face attendance workflow should complete", "Medium"),
        ("Faculty", "Marks", "Marks Entry", "TC-FAC-007", "Verify faculty can enter marks", "Open Enter Marks and save mark data", "Marks entry should save successfully", "High"),
        ("Faculty", "Assignments", "Assignment Management", "TC-FAC-008", "Verify assignments workflow", "Open Assignments page and create or manage assignment records", "Assignment workflow should work", "Medium"),
        ("Faculty", "Result Upload", "Result Upload", "TC-FAC-009", "Verify faculty can upload results", "Open Result Upload and submit result data", "Result upload should complete", "High"),
        ("Faculty", "Study Materials", "Material Upload", "TC-FAC-010", "Verify faculty can upload materials", "Open Study Materials page and upload content", "Material should be saved and visible", "Medium"),
        ("Faculty", "Students", "Faculty Student View", "TC-FAC-011", "Verify faculty can view assigned students", "Open My Students page", "Assigned student list should appear", "Medium"),
        ("Faculty", "Leave", "Faculty Leave Request", "TC-FAC-012", "Verify faculty can submit leave request", "Open Leave Request page and submit leave form", "Leave request should be created", "Medium"),
        ("Faculty", "Payslips", "Faculty Payslips", "TC-FAC-013", "Verify faculty can view payslips", "Open My Payslips page", "Payslip or salary details should load", "Low"),
        ("HOD", "Dashboard", "HOD Dashboard", "TC-HOD-001", "Verify HOD dashboard loads", "Login as HOD and open dashboard", "HOD dashboard should load", "High"),
        ("HOD", "Faculty Management", "Department Faculty", "TC-HOD-002", "Verify department faculty page loads", "Open Dept Faculty page", "Department faculty list should load", "High"),
        ("HOD", "Department Overview", "Department Overview", "TC-HOD-003", "Verify department overview loads", "Open Dept Overview page", "Department overview data should load", "High"),
        ("HOD", "Courses", "Department Courses", "TC-HOD-004", "Verify HOD can manage courses", "Open Courses page and review or update course-related data", "Course management should work", "High"),
        ("HOD", "Timetable", "Department Timetable", "TC-HOD-005", "Verify HOD can manage timetable", "Open Timetable page and manage scheduling data", "Department timetable workflow should work", "High"),
        ("HOD", "Leave Approval", "Leave Approval", "TC-HOD-006", "Verify HOD can approve leave", "Open Leave Approvals and approve or reject request", "Leave approval action should save", "High"),
        ("HOD", "Notifications", "HOD Notifications", "TC-HOD-007", "Verify HOD notifications page loads", "Open Notifications page", "Department notifications should load", "Medium"),
        ("Admin", "Dashboard", "Admin Dashboard", "TC-ADM-001", "Verify admin dashboard loads", "Login as admin and open dashboard", "Summary cards, charts, recent activities, and quick actions should load", "High"),
        ("Admin", "Students", "Student Management", "TC-ADM-002", "Verify student management page loads", "Open Students page", "Student management table and filters should load", "High"),
        ("Admin", "Students", "Student Creation", "TC-ADM-003", "Verify admin can add student", "Use Add student and save a new record", "Student should be created", "High"),
        ("Admin", "Students", "Student Update", "TC-ADM-004", "Verify admin can edit student", "Open existing student record and save changes", "Changes should be saved", "High"),
        ("Admin", "Students", "Student Deletion", "TC-ADM-005", "Verify admin can delete student", "Delete a student and confirm action", "Student should be removed", "High"),
        ("Admin", "Students", "Student Export Excel", "TC-ADM-006", "Verify Excel export from student management", "Use Excel export in student management", "Excel export should download", "Medium"),
        ("Admin", "Students", "Student Export CSV", "TC-ADM-007", "Verify CSV export from student management", "Use CSV export in student management", "CSV export should download", "Medium"),
        ("Admin", "Faculty", "Faculty Management", "TC-ADM-008", "Verify faculty management workflow", "Open Faculty page and perform CRUD checks", "Faculty workflow should work", "High"),
        ("Admin", "Departments", "Department Management", "TC-ADM-009", "Verify department management workflow", "Open Departments page and perform CRUD checks", "Department workflow should work", "High"),
        ("Admin", "Courses", "Course Management", "TC-ADM-010", "Verify course management workflow", "Open Curriculum or Courses page and perform CRUD checks", "Course workflow should work", "High"),
        ("Admin", "Admissions", "Admission Management", "TC-ADM-011", "Verify admission management workflow", "Open Admissions page and review or update applications", "Admission workflow should work", "High"),
        ("Admin", "Attendance", "Attendance Review", "TC-ADM-012", "Verify attendance management view loads", "Open Attendance page", "Attendance management view should load", "Medium"),
        ("Admin", "Exams", "Exam Management", "TC-ADM-013", "Verify exam workflow", "Open Exams page and perform exam operations", "Exam workflow should work", "High"),
        ("Admin", "Results", "Result Management", "TC-ADM-014", "Verify result management workflow", "Open Results page and review or update result data", "Result management should work", "High"),
        ("Admin", "Fees", "Fee Management", "TC-ADM-015", "Verify fee management workflow", "Open Fees page and perform fee operations", "Fee workflow should work", "High"),
        ("Admin", "Timetable", "Timetable Management", "TC-ADM-016", "Verify timetable management workflow", "Open Timetable page and perform scheduling operations", "Timetable workflow should work", "High"),
        ("Admin", "Notifications", "Notice Management", "TC-ADM-017", "Verify admin can create notice", "Open Notifications page and create a notice", "Notice should be published", "High"),
        ("Admin", "Library", "Library Management", "TC-ADM-018", "Verify library workflow", "Open Library page and perform book issue or return checks", "Library workflow should work", "Medium"),
        ("Admin", "Placements", "Placement Management", "TC-ADM-019", "Verify placement workflow", "Open Placements page and perform placement operations", "Placement workflow should work", "Medium"),
        ("Admin", "Certificates", "Certificate Management", "TC-ADM-020", "Verify certificate workflow", "Open Certificates page and generate or review certificate records", "Certificate workflow should work", "Medium"),
        ("Admin", "Reports", "Reports", "TC-ADM-021", "Verify reports page loads", "Open Reports page", "Analytics or report data should load", "Medium"),
        ("Admin", "Hostel", "Hostel Management", "TC-ADM-022", "Verify hostel workflow", "Open Hostel page", "Hostel management workflow should work", "Low"),
        ("Admin", "Transport", "Transport Management", "TC-ADM-023", "Verify transport workflow", "Open Transport page", "Transport workflow should work", "Low"),
        ("Admin", "Leave Approvals", "Leave Approval", "TC-ADM-024", "Verify admin leave approvals workflow", "Open Leave Approvals page and process requests", "Leave action should save", "Medium"),
        ("Admin", "Payroll", "Payroll Management", "TC-ADM-025", "Verify payroll workflow", "Open Payroll page", "Payroll workflow should load", "Medium"),
        ("Admin", "Scholarship", "Scholarship Management", "TC-ADM-026", "Verify scholarship workflow", "Open Scholarship page", "Scholarship workflow should load", "Low"),
        ("Admin", "Events", "Event Management", "TC-ADM-027", "Verify event approval workflow", "Open Events page", "Event approval workflow should load", "Low"),
        ("Admin", "Result Upload", "Admin Result Upload", "TC-ADM-028", "Verify admin result upload workflow", "Open Result Upload and submit result data", "Admin result upload should work", "Medium"),
        ("Admin", "Profile", "Admin Profile", "TC-ADM-029", "Verify admin profile page loads", "Open My Profile page", "Admin profile page should load", "Low"),
    ]


def style_sheet(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_overview_sheet(wb, rows):
    ws = wb.active
    ws.title = "Overview"

    ws["A1"] = "College Management System Website Test Cases And Test Scenarios"
    ws["A1"].font = Font(size=16, bold=True, color="173A5E")
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Generated on {date.today().strftime('%d %B %Y')}"
    ws.merge_cells("A2:F2")
    ws["A3"] = "Use this workbook to execute role-based website test cases and record results."
    ws.merge_cells("A3:F3")

    for row_ref in ("A1", "A2", "A3"):
        ws[row_ref].fill = TITLE_FILL
        ws[row_ref].alignment = Alignment(horizontal="center")

    headers = ["Sheet", "Purpose", "How to use", "Total Test Cases", "Owner", "Status Values"]
    data = [
        ("Master Test Cases", "Complete set of website test cases", "Execute each row and update result columns", len(rows), "QA / Project Team", "Not Started, In Progress, Pass, Fail, Blocked, N/A"),
        ("Public Visitor", "Test scenarios for non-logged-in users", "Review landing and registration flows", "", "QA / Demo User", ""),
        ("Student", "Test scenarios for student workflows", "Review student portal pages and actions", "", "QA / Student User", ""),
        ("Faculty", "Test scenarios for faculty workflows", "Review faculty teaching and academic tasks", "", "QA / Faculty User", ""),
        ("HOD", "Test scenarios for HOD workflows", "Review department management and leave approval", "", "QA / HOD User", ""),
        ("Admin", "Test scenarios for admin workflows", "Review institution-wide management flows", "", "QA / Admin User", ""),
    ]

    start_row = 5
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, values in enumerate(data, start=start_row + 1):
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_index, col, value)
            if row_index % 2 == 0:
                cell.fill = ALT_FILL

    ws.freeze_panes = "A5"
    style_sheet(ws, [24, 30, 34, 16, 20, 28])


def add_test_sheet(wb, title, rows):
    ws = wb.create_sheet(title)
    headers = [
        "S.No",
        "Role",
        "Module",
        "Test Scenario",
        "Test Case ID",
        "Test Case",
        "Test Steps",
        "Expected Result",
        "Priority",
        "Status",
        "Tester",
        "Test Date",
        "Remarks",
    ]

    ws["A1"] = title
    ws.merge_cells("A1:M1")
    ws["A1"].font = Font(size=14, bold=True, color="173A5E")
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    header_row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, header)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    current_module = None
    row_index = header_row + 1
    serial = 1
    for role, module, scenario, case_id, test_case, steps, expected, priority in rows:
        if module != current_module:
            for col in range(1, len(headers) + 1):
                sec = ws.cell(row_index, col)
                if col == 1:
                    sec.value = f"{module} Section"
                sec.fill = SECTION_FILL
                sec.font = Font(bold=True, color="1D3557")
            row_index += 1
            current_module = module

        values = [serial, role, module, scenario, case_id, test_case, steps, expected, priority, "", "", "", ""]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_index, col, value)
            if serial % 2 == 0:
                cell.fill = ALT_FILL
        row_index += 1
        serial += 1

    status_validation = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Pass,Fail,Blocked,N/A"',
        allow_blank=True,
    )
    ws.add_data_validation(status_validation)
    status_validation.add(f"J{header_row + 1}:J{row_index}")

    ws.freeze_panes = f"A{header_row + 1}"
    style_sheet(ws, [8, 14, 18, 22, 14, 28, 34, 34, 11, 15, 18, 14, 28])


def build_workbook():
    rows = test_rows()
    wb = Workbook()
    build_overview_sheet(wb, rows)
    add_test_sheet(wb, "Master Test Cases", rows)

    for role_name in ["Public Visitor", "Student", "Faculty", "HOD", "Admin"]:
        add_test_sheet(wb, role_name, [row for row in rows if row[0] == role_name])

    wb.save(OUTPUT_FILE)
    print(f"Generated: {OUTPUT_FILE}")
    print(f"Total test cases: {len(rows)}")


if __name__ == "__main__":
    build_workbook()
